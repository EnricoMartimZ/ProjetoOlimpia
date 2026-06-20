"""
Gera a planilha (casos_de_uso.xlsx) que sintetiza os casos de uso e cenários
de teste do Projeto Olímpia. Saída na mesma pasta (use_cases/).

Uso: python _gerar_planilha.py
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Dados sintetizados
# ---------------------------------------------------------------------------

# Casos de uso: (id, nome, requisitos, ator, pre, pos, regras, rotas, qtd_cenarios)
CASOS = [
    ("UC01", "Autenticar no sistema (Login)", "REQ 2, 3",
     "Servidor / Pesquisador de campo",
     "Usuário cadastrado; aplicação disponível.",
     "Em sucesso, devolve JWT (sub, nome, role, exp) para acessar rotas protegidas.",
     "Erro de credencial genérico (401); senha só em hash; token expira; role define autorização.",
     "POST /auth/login", 6),
    ("UC02", "Cadastrar usuário", "REQ 4",
     "Servidor da Secretaria",
     "E-mail não cadastrado; ao menos um role válido.",
     "Usuário persistido com senha hasheada; passa a poder logar.",
     "E-mail único; senha só hash (nunca exposta); >=1 role; roles deduplicados; roles ∈ {servidor, pesquisador_campo}.",
     "POST /usuarios", 8),
    ("UC03", "Criar e gerenciar pesquisa", "REQ 5",
     "Servidor da Secretaria",
     "Servidor autenticado; nome único.",
     "Pesquisa criada com campos base e tipo; nasce em rascunho.",
     "Nome único; tipo publica/campo (default publica); status derivado; PUT substitui campos; delete em cascata.",
     "POST/GET/PUT/DELETE /pesquisas", 11),
    ("UC04", "Lançar edição de pesquisa", "REQ 6",
     "Servidor da Secretaria",
     "Pesquisa existe; servidor autenticado.",
     "Edição criada com numero_edicao auto-incrementado; status derivado do período.",
     "numero_edicao auto; status agendada/ativa/encerrada; fechamento >= abertura; campos = fixos + extras.",
     "POST /pesquisas/{id}/edicoes; GET .../edicoes; GET /edicoes/{id}/campos", 10),
    ("UC05", "Responder pesquisa pública", "REQ 8, 9",
     "Cidadão / Dono de hospedagem",
     "Edição de pesquisa publica e aberta (ativa).",
     "Resposta persistida; grava usuario_id se autenticado, senão null.",
     "Só tipo publica; edição ativa; campo pertence à edição, sem duplicata; lista não vazia; auth opcional.",
     "GET /publico/edicoes/{id}; POST /edicoes/{id}/respostas", 10),
    ("UC06", "Coletar resposta de campo", "REQ 3, 10",
     "Pesquisador de campo",
     "Pesquisa tipo campo com edição aberta; pesquisador autenticado.",
     "Resposta vinculada ao usuario_id do pesquisador e à edição.",
     "Só role pesquisador_campo; sempre vincula coletor; só edições tipo campo; lista só abertas; validações de resposta.",
     "GET/POST /pesquisador/edicoes[/ {id}[/respostas]]", 9),
    ("UC07", "Visualizar respostas tabuladas", "REQ 11",
     "Servidor da Secretaria",
     "Servidor autenticado; edição existe.",
     "Respostas tabuladas, paginadas, com cabeçalho; remoção apaga resposta.",
     "Só servidor; paginação por_pagina 1-500 (def 20); usuario_nome = coletor (null se anônimo); busca ILIKE; rid deve ser da edição.",
     "GET/DELETE /edicoes/{id}/respostas[/{rid}]", 10),
    ("UC08", "Gerenciar hospedagens (Diária Média)", "REQ 7",
     "Servidor da Secretaria",
     "Servidor autenticado; CNPJ não cadastrado (criar).",
     "Hospedagem persistida; exclusão em cascata nas diárias.",
     "CNPJ chave/imutável e pode ter '/' (cru na URL); estrelas 0-5; quartos>=0; categoria default Hotel; delete cascata.",
     "CRUD /hospedagens", 10),
    ("UC09", "Registrar diária média", "REQ 7",
     "Servidor da Secretaria",
     "Servidor autenticado; hospedagem existe; sem registro na data.",
     "Registro (data+preco) persistido; hospedagem sai de pendentes na data.",
     "Só servidor; 1 registro por hospedagem+data (409 duplicado); preco>=0; hospedagem deve existir; pendentes = sem diária na data.",
     "POST/GET/DELETE /diarias; GET /diarias/pendentes", 10),
]

# Cenários: (id, caso, tipo, titulo, pre, passos, esperado)
CENARIOS = [
    # UC01
    ("UC01-CT01", "UC01", "Principal", "Login de servidor com credenciais válidas",
     "Servidor cadastrado", "POST /auth/login com email+senha válidos", "200; access_token + token_type=bearer"),
    ("UC01-CT02", "UC01", "Alternativo", "Login de pesquisador de campo",
     "Pesquisador cadastrado", "POST /auth/login com credenciais do pesquisador", "200; token com role=pesquisador_campo"),
    ("UC01-CT03", "UC01", "Exceção", "E-mail inexistente",
     "—", "POST /auth/login com e-mail não cadastrado", "401 (credenciais inválidas, genérico)"),
    ("UC01-CT04", "UC01", "Exceção", "Senha incorreta",
     "Servidor cadastrado", "POST /auth/login com senha errada", "401; idêntico a e-mail inexistente"),
    ("UC01-CT05", "UC01", "Exceção", "Payload incompleto/inválido",
     "—", "POST /auth/login sem senha / corpo vazio", "422"),
    ("UC01-CT06", "UC01", "Alternativo", "Conteúdo do token (claims)",
     "Servidor logado", "Decodificar access_token", "Claims sub, nome, role=servidor, exp futuro"),
    # UC02
    ("UC02-CT01", "UC02", "Principal", "Cadastrar servidor com dados válidos",
     "—", "POST /usuarios com roles=[servidor]", "201; sem senha; roles=[servidor]; criado_em"),
    ("UC02-CT02", "UC02", "Alternativo", "Cadastrar pesquisador de campo",
     "—", "POST /usuarios com roles=[pesquisador_campo]", "201; roles=[pesquisador_campo]"),
    ("UC02-CT03", "UC02", "Alternativo", "Múltiplos papéis com duplicata",
     "—", "POST /usuarios roles=[servidor,servidor,pesquisador_campo]", "201; roles deduplicado e ordenado"),
    ("UC02-CT04", "UC02", "Exceção", "E-mail já cadastrado",
     "Usuário com e-mail existe", "POST /usuarios com mesmo e-mail", "409"),
    ("UC02-CT05", "UC02", "Exceção", "Lista de roles vazia",
     "—", "POST /usuarios com roles=[]", "422"),
    ("UC02-CT06", "UC02", "Exceção", "Role inválido",
     "—", "POST /usuarios com role inexistente", "422"),
    ("UC02-CT07", "UC02", "Exceção", "E-mail inválido / campo faltando",
     "—", "POST /usuarios e-mail inválido / sem senha", "422"),
    ("UC02-CT08", "UC02", "Alternativo", "Senha hasheada e não exposta",
     "Usuário criado", "Inspecionar registro no banco; logar", "Senha no banco != original; login funciona"),
    # UC03
    ("UC03-CT01", "UC03", "Principal", "Criar pesquisa com campos",
     "Servidor autenticado", "POST /pesquisas com 2 campos", "201; status=rascunho; total_edicoes=0; campos com id/hash/ordem"),
    ("UC03-CT02", "UC03", "Alternativo", "Tipo omitido assume publica",
     "—", "POST /pesquisas sem tipo", "201; tipo=publica"),
    ("UC03-CT03", "UC03", "Alternativo", "Criar pesquisa tipo campo",
     "—", "POST /pesquisas tipo=campo", "201; tipo=campo"),
    ("UC03-CT04", "UC03", "Alternativo", "Detalhar pesquisa com campos",
     "Pesquisa criada", "GET /pesquisas/{id}", "200; campos completos"),
    ("UC03-CT05", "UC03", "Alternativo", "Listar pesquisas (derivados)",
     "≥1 pesquisa", "GET /pesquisas", "200; status, total_edicoes, edicao_atual_id, tipo"),
    ("UC03-CT06", "UC03", "Alternativo", "Editar substituindo campos",
     "Pesquisa com 2 campos", "PUT /pesquisas/{id} com 1 campo novo", "200; só o novo campo (substitui)"),
    ("UC03-CT07", "UC03", "Alternativo", "Excluir em cascata",
     "Pesquisa com edição+respostas", "DELETE /pesquisas/{id}", "204; GET posterior 404; cascata"),
    ("UC03-CT08", "UC03", "Exceção", "Criar sem ser servidor",
     "—", "POST /pesquisas sem token / como pesquisador", "401 / 403"),
    ("UC03-CT09", "UC03", "Exceção", "Nome duplicado",
     "Pesquisa com nome X existe", "POST /pesquisas com mesmo nome", "409"),
    ("UC03-CT10", "UC03", "Exceção", "Pesquisa inexistente",
     "—", "GET/PUT/DELETE /pesquisas/999999", "404"),
    ("UC03-CT11", "UC03", "Exceção", "Tipo de campo inválido",
     "—", "POST /pesquisas com campo de tipo inválido", "422"),
    # UC04
    ("UC04-CT01", "UC04", "Principal", "Lançar primeira edição",
     "Pesquisa criada", "POST /pesquisas/{id}/edicoes abertura=hoje", "201; numero_edicao=1; status=ativa; total_respostas=0"),
    ("UC04-CT02", "UC04", "Alternativo", "Segunda edição incrementa número",
     "Pesquisa com 1 edição", "POST /pesquisas/{id}/edicoes", "201; numero_edicao=2"),
    ("UC04-CT03", "UC04", "Alternativo", "Edição agendada",
     "—", "POST abertura no futuro", "201; status=agendada"),
    ("UC04-CT04", "UC04", "Alternativo", "Edição sem fechamento fica ativa",
     "—", "POST abertura=hoje sem fechamento", "201; status=ativa"),
    ("UC04-CT05", "UC04", "Alternativo", "Campos extras e combinação ordenada",
     "Pesquisa com 1 campo fixo", "POST com campos_extras; GET /edicoes/{id}/campos", "200; fixo antes do extra"),
    ("UC04-CT06", "UC04", "Alternativo", "Listar edições",
     "Pesquisa com 2 edições", "GET /pesquisas/{id}/edicoes", "200; 2 itens com status/total_respostas"),
    ("UC04-CT07", "UC04", "Exceção", "Lançar sem ser servidor",
     "—", "POST sem token / como pesquisador", "401 / 403"),
    ("UC04-CT08", "UC04", "Exceção", "Pesquisa inexistente",
     "—", "POST /pesquisas/999999/edicoes", "404"),
    ("UC04-CT09", "UC04", "Exceção", "Datas inconsistentes",
     "—", "POST fechamento < abertura", "422"),
    ("UC04-CT10", "UC04", "Exceção", "Campos de edição inexistente",
     "—", "GET /edicoes/999999/campos", "404"),
    # UC05
    ("UC05-CT01", "UC05", "Principal", "Abrir formulário e enviar resposta anônima",
     "Pesquisa publica, edição ativa", "GET /publico/edicoes/{id}; POST /edicoes/{id}/respostas", "200 (aberta=true); 201; usuario_id=null"),
    ("UC05-CT02", "UC05", "Alternativo", "Envio autenticado grava usuario_id",
     "Edição ativa; usuário logado", "POST /edicoes/{id}/respostas com token", "201; usuario_id do token"),
    ("UC05-CT03", "UC05", "Alternativo", "Formulário fora do período",
     "Edição agendada/encerrada", "GET /publico/edicoes/{id}", "200; aberta=false"),
    ("UC05-CT04", "UC05", "Exceção", "Formulário público de edição de campo",
     "Pesquisa tipo campo", "GET /publico/edicoes/{id}", "404"),
    ("UC05-CT05", "UC05", "Exceção", "Envio em edição de campo",
     "Edição de pesquisa campo", "POST /edicoes/{id}/respostas", "403"),
    ("UC05-CT06", "UC05", "Exceção", "Envio em edição encerrada",
     "Edição publica encerrada", "POST /edicoes/{id}/respostas", "409"),
    ("UC05-CT07", "UC05", "Exceção", "campo_id de outra edição",
     "Edições A e B", "POST /edicoes/A/respostas com campo de B", "422"),
    ("UC05-CT08", "UC05", "Exceção", "campo_id duplicado",
     "—", "POST com 2 itens de mesmo campo_id", "422"),
    ("UC05-CT09", "UC05", "Exceção", "Lista de respostas vazia",
     "—", "POST com respostas=[]", "422"),
    ("UC05-CT10", "UC05", "Exceção", "Edição inexistente",
     "—", "POST/GET em edição 999999", "404"),
    # UC06
    ("UC06-CT01", "UC06", "Principal", "Coleta vinculada ao pesquisador",
     "Pesquisa campo, edição ativa; pesquisador", "GET edicoes; GET edicoes/{id}; POST respostas", "200/200/201; resposta com usuario_id do pesquisador"),
    ("UC06-CT02", "UC06", "Alternativo", "Sem edições de campo abertas",
     "Nenhuma edição campo aberta", "GET /pesquisador/edicoes", "200; lista vazia"),
    ("UC06-CT03", "UC06", "Alternativo", "Múltiplas coletas na mesma edição",
     "Edição campo ativa", "POST respostas 3x", "3x 201; todas com usuario_id do pesquisador"),
    ("UC06-CT04", "UC06", "Exceção", "Servidor na rota de pesquisador",
     "—", "GET /pesquisador/edicoes como servidor", "403"),
    ("UC06-CT05", "UC06", "Exceção", "Anônimo na rota de pesquisador",
     "—", "GET /pesquisador/edicoes sem token", "401"),
    ("UC06-CT06", "UC06", "Exceção", "Edição pública via fluxo de campo",
     "Edição de pesquisa publica", "GET/POST /pesquisador/edicoes/{id}", "404"),
    ("UC06-CT07", "UC06", "Exceção", "Coleta em edição fechada",
     "Edição campo encerrada", "POST /pesquisador/edicoes/{id}/respostas", "409"),
    ("UC06-CT08", "UC06", "Exceção", "Campo inválido/duplicado/lista vazia",
     "—", "POST com campo de outra edição / duplicado / vazio", "422"),
    ("UC06-CT09", "UC06", "Exceção", "Edição inexistente",
     "—", "GET/POST /pesquisador/edicoes/999999", "404"),
    # UC07
    ("UC07-CT01", "UC07", "Principal", "Consultar respostas tabuladas",
     "Edição com ≥1 resposta; servidor", "GET /edicoes/{id}/respostas", "200; total≥1; campos_header; valores; anônimo usuario_nome=null"),
    ("UC07-CT02", "UC07", "Alternativo", "Paginação",
     "Edição com 3 respostas", "GET ?pagina=1&por_pagina=2", "200; len(dados)=2; total=3"),
    ("UC07-CT03", "UC07", "Alternativo", "Busca por texto (ILIKE)",
     "Respostas com textos diferentes", "GET ?busca=são", "200; só linhas que casam (case-insensitive)"),
    ("UC07-CT04", "UC07", "Alternativo", "Edição sem respostas",
     "Edição com 0 respostas", "GET /edicoes/{id}/respostas", "200; total=0; dados=[]; header preenchido"),
    ("UC07-CT05", "UC07", "Alternativo", "Coletado por em coleta de campo",
     "Resposta coletada por pesquisador", "GET /edicoes/{id}/respostas", "usuario_nome = nome do pesquisador"),
    ("UC07-CT06", "UC07", "Alternativo", "Remover resposta",
     "Edição com resposta rid", "DELETE /edicoes/{id}/respostas/{rid}", "204; some da consulta; total decrementa"),
    ("UC07-CT07", "UC07", "Exceção", "Consultar sem ser servidor",
     "—", "GET sem token / como pesquisador", "401 / 403"),
    ("UC07-CT08", "UC07", "Exceção", "Edição inexistente",
     "—", "GET /edicoes/999999/respostas", "404"),
    ("UC07-CT09", "UC07", "Exceção", "Remover resposta de outra edição",
     "rid pertence à edição B", "DELETE /edicoes/A/respostas/{rid}", "404"),
    ("UC07-CT10", "UC07", "Exceção", "Paginação inválida",
     "—", "GET ?pagina=0 / ?por_pagina=1000", "422"),
    # UC08
    ("UC08-CT01", "UC08", "Principal", "Cadastrar hospedagem",
     "Servidor autenticado", "POST /hospedagens payload completo", "201; cnpj/nome/url_booking; criado_em"),
    ("UC08-CT02", "UC08", "Alternativo", "Categoria default",
     "—", "POST /hospedagens sem categoria", "201; categoria=Hotel"),
    ("UC08-CT03", "UC08", "Alternativo", "Listar ordenado por nome",
     "2 hospedagens", "GET /hospedagens", "200; ordem alfabética por nome"),
    ("UC08-CT04", "UC08", "Alternativo", "Detalhar por CNPJ com barra",
     "Hospedagem com CNPJ", "GET /hospedagens/12.345.678/0001-90", "200; hospedagem correta"),
    ("UC08-CT05", "UC08", "Alternativo", "Editar parcial (CNPJ imutável)",
     "Hospedagem cadastrada", "PUT /hospedagens/{cnpj} {estrelas:4}", "200; estrelas=4; cnpj inalterado"),
    ("UC08-CT06", "UC08", "Alternativo", "Excluir com cascata",
     "Hospedagem com ≥1 diária", "DELETE /hospedagens/{cnpj}", "204; GET 404; diárias removidas"),
    ("UC08-CT07", "UC08", "Exceção", "Cadastrar sem ser servidor",
     "—", "POST sem token / como pesquisador", "401 / 403"),
    ("UC08-CT08", "UC08", "Exceção", "CNPJ duplicado",
     "CNPJ já cadastrado", "POST /hospedagens mesmo CNPJ", "409"),
    ("UC08-CT09", "UC08", "Exceção", "Estrelas/quartos inválidos",
     "—", "POST estrelas=7 / quartos=-1", "422"),
    ("UC08-CT10", "UC08", "Exceção", "Hospedagem inexistente",
     "—", "GET/PUT/DELETE /hospedagens/{cnpj inexistente}", "404"),
    # UC09
    ("UC09-CT01", "UC09", "Principal", "Registrar diária",
     "Hospedagem cadastrada; servidor", "POST /diarias {cnpj, data=hoje, preco}", "201; id, nome_fantasia, data, preco, registrado_em"),
    ("UC09-CT02", "UC09", "Alternativo", "Pendentes deixa de listar após registro",
     "Hospedagem pendente hoje", "GET pendentes; POST /diarias; GET pendentes", "Após registro, some dos pendentes de hoje"),
    ("UC09-CT03", "UC09", "Alternativo", "Listar com filtros",
     "≥2 registros", "GET /diarias [?hospedagem_cnpj / ?data]", "200; filtrado; mais recentes primeiro"),
    ("UC09-CT04", "UC09", "Alternativo", "Mesma hospedagem em datas diferentes",
     "—", "POST /diarias mesmo cnpj, hoje e hoje-1", "ambos 201 (unicidade por hospedagem+data)"),
    ("UC09-CT05", "UC09", "Alternativo", "Remover registro e voltar a pendente",
     "Diária id registrada hoje", "DELETE /diarias/{id}; GET pendentes", "204; hospedagem volta a pendente"),
    ("UC09-CT06", "UC09", "Exceção", "Registrar sem ser servidor",
     "—", "POST /diarias sem token / como pesquisador", "401 / 403"),
    ("UC09-CT07", "UC09", "Exceção", "Hospedagem inexistente",
     "—", "POST /diarias cnpj não cadastrado", "404"),
    ("UC09-CT08", "UC09", "Exceção", "Registro duplicado hospedagem+data",
     "Já existe diária na data X", "POST /diarias mesma hospedagem+data X", "409"),
    ("UC09-CT09", "UC09", "Exceção", "Preço negativo",
     "—", "POST /diarias preco=-10", "422"),
    ("UC09-CT10", "UC09", "Exceção", "Remover registro inexistente",
     "—", "DELETE /diarias/999999", "404"),
]

# ---------------------------------------------------------------------------
# Estilo
# ---------------------------------------------------------------------------
AZUL = "1F4E78"
AZUL_CLARO = "DDEBF7"
CINZA = "F2F2F2"
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor=AZUL)
title_font = Font(bold=True, size=16, color=AZUL)
wrap = Alignment(wrap_text=True, vertical="top")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

TIPO_COR = {
    "Principal": "C6EFCE",
    "Alternativo": "FFF2CC",
    "Exceção": "FCE4D6",
}


def estiliza_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border


def aplica_bordas_zebra(ws, first_row, last_row, ncols, zebra=True):
    for r in range(first_row, last_row + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = wrap
            cell.border = border
        if zebra and (r - first_row) % 2 == 1:
            for c in range(1, ncols + 1):
                ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=CINZA)


# Fluxos resumidos por caso (principal / alternativos / exceção)
FLUXOS = {
    "UC01": ("Envia credenciais; valida usuário/senha; gera JWT; retorna 200.",
             "A1 pesquisador (role no token); A2 validação de aba no front.",
             "E1 e-mail inexistente 401; E2 senha errada 401; E3 payload inválido 422."),
    "UC02": ("Envia dados; valida; checa e-mail; hasheia senha; persiste; 201.",
             "A1 múltiplos papéis (dedup); A2 cadastrar pesquisador.",
             "E1 e-mail duplicado 409; E2 roles vazio 422; E3 role inválido 422; E4 e-mail/campos 422."),
    "UC03": ("Envia pesquisa+campos; valida; nome único; persiste; 201 (rascunho).",
             "A1 tipo default publica; A2 tipo campo; A3 detalhar; A4 listar; A5 editar (substitui); A6 excluir cascata.",
             "E1 não-servidor 401/403; E2 nome duplicado 409; E3 inexistente 404; E4 tipo de campo inválido 422."),
    "UC04": ("Envia edição; confere pesquisa; auto-incrementa número; persiste; 201.",
             "A1 sem fechamento; A2 agendada; A3 sem extras; A4 listar; A5 campos combinados; A6 2ª edição.",
             "E1 não-servidor 401/403; E2 pesquisa inexistente 404; E3 datas 422; E4 campos de edição inexistente 404."),
    "UC05": ("Abre formulário; envia respostas; valida campos/edição aberta; persiste; 201.",
             "A1 autenticado grava usuario_id; A2 edição fora do período (aberta=false).",
             "E1 edição de campo 404/403; E2 fechada 409; E3 campo inválido 422; E4 duplicado 422; E5 vazia 422; E6 inexistente 404."),
    "UC06": ("Loga; lista edições abertas; abre form; envia coleta vinculada; 201.",
             "A1 nenhuma edição aberta (lista vazia); A2 múltiplas coletas.",
             "E1 servidor 403; E2 anônimo 401; E3 não é campo 404; E4 fechada 409; E5 campo/lista 422; E6 inexistente 404."),
    "UC07": ("Consulta; monta cabeçalho; retorna tabela paginada com valores.",
             "A1 paginação; A2 busca ILIKE; A3 sem respostas; A4 coletado por; A5 remover (204).",
             "E1 não-servidor 401/403; E2 edição inexistente 404; E3 rid de outra edição 404; E4 paginação inválida 422."),
    "UC08": ("Envia hospedagem; valida; CNPJ único; persiste; 201.",
             "A1 categoria default Hotel; A2 listar; A3 detalhar; A4 editar parcial; A5 excluir cascata.",
             "E1 não-servidor 401/403; E2 CNPJ duplicado 409; E3 validação 422; E4 inexistente 404."),
    "UC09": ("Consulta pendentes; envia diária; valida; persiste; 201.",
             "A1 data default hoje; A2 listar com filtros; A3 datas diferentes; A4 remover (204).",
             "E1 não-servidor 401/403; E2 hospedagem inexistente 404; E3 duplicado 409; E4 preço negativo 422; E5 remover inexistente 404."),
}

wb = Workbook()

# ===== Aba 1: Casos de Uso (visão completa) =====
ws = wb.active
ws.title = "Casos de Uso"
ws["A1"] = "Projeto Olímpia (MERX) — Casos de Uso"
ws["A1"].font = title_font
ws.merge_cells("A1:L1")
ws["A2"] = "Sistema de Gestão de Pesquisas para o Turismo de Olímpia — Secretaria de Turismo"
ws["A2"].font = Font(italic=True, color="595959")
ws.merge_cells("A2:L2")

cab = ["ID", "Caso de uso", "Requisito(s)", "Ator", "Pré-condições", "Pós-condições",
       "Fluxo principal", "Fluxos alternativos", "Fluxos de exceção",
       "Regras de negócio", "Rotas envolvidas", "Nº cenários"]
hrow = 4
for i, h in enumerate(cab, 1):
    ws.cell(row=hrow, column=i, value=h)
estiliza_header(ws, hrow, len(cab))

r = hrow + 1
for uc in CASOS:
    cid, nome, req, ator, pre, pos, regras, rotas, qtd = uc
    fp, fa, fe = FLUXOS[cid]
    vals = [cid, nome, req, ator, pre, pos, fp, fa, fe, regras, rotas, qtd]
    for i, v in enumerate(vals, 1):
        cell = ws.cell(row=r, column=i, value=v)
        if i == 12:
            cell.alignment = center
    r += 1
aplica_bordas_zebra(ws, hrow + 1, r - 1, len(cab))

# linha total
ws.cell(row=r, column=1, value="TOTAL").font = Font(bold=True)
tot_cell = ws.cell(row=r, column=12, value=sum(uc[8] for uc in CASOS))
tot_cell.font = Font(bold=True)
tot_cell.alignment = center
for c in range(1, len(cab) + 1):
    ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=AZUL_CLARO)
    ws.cell(row=r, column=c).border = border

larguras = [8, 28, 12, 22, 30, 32, 40, 44, 48, 44, 32, 11]
for i, w in enumerate(larguras, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "C5"

# ===== Aba 2: Cenários de Teste =====
ws3 = wb.create_sheet("Cenários de Teste")
ws3["A1"] = "Cenários de Teste"
ws3["A1"].font = title_font
ws3.merge_cells("A1:H1")
cab3 = ["ID Cenário", "Caso de uso", "Tipo de fluxo", "Título", "Pré-condições", "Passos", "Resultado esperado", "OK?"]
hrow3 = 3
for i, h in enumerate(cab3, 1):
    ws3.cell(row=hrow3, column=i, value=h)
estiliza_header(ws3, hrow3, len(cab3))
r = hrow3 + 1
for cen in CENARIOS:
    cid, caso, tipo, titulo, pre, passos, esperado = cen
    row_vals = [cid, caso, tipo, titulo, pre, passos, esperado, ""]
    for i, v in enumerate(row_vals, 1):
        ws3.cell(row=r, column=i, value=v)
    # cor por tipo
    cor = TIPO_COR.get(tipo)
    if cor:
        ws3.cell(row=r, column=3).fill = PatternFill("solid", fgColor=cor)
    r += 1
# bordas (sem zebra para não conflitar com cor de tipo)
for rr in range(hrow3 + 1, r):
    for c in range(1, len(cab3) + 1):
        cell = ws3.cell(row=rr, column=c)
        cell.alignment = wrap
        cell.border = border
for i, w in enumerate([14, 12, 14, 38, 30, 44, 46, 8], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w
ws3.freeze_panes = "A4"
ws3.auto_filter.ref = f"A{hrow3}:H{r-1}"

# ===== Aba 3: Resumo / contagem =====
ws4 = wb.create_sheet("Resumo")
ws4["A1"] = "Resumo dos cenários por tipo de fluxo"
ws4["A1"].font = title_font
ws4.merge_cells("A1:E1")
cab4 = ["Caso de uso", "Principal", "Alternativo", "Exceção", "Total"]
hrow4 = 3
for i, h in enumerate(cab4, 1):
    ws4.cell(row=hrow4, column=i, value=h)
estiliza_header(ws4, hrow4, len(cab4))
r = hrow4 + 1
tot_p = tot_a = tot_e = 0
for uc in CASOS:
    cens = [c for c in CENARIOS if c[1] == uc[0]]
    p = sum(1 for c in cens if c[2] == "Principal")
    a = sum(1 for c in cens if c[2] == "Alternativo")
    e = sum(1 for c in cens if c[2] == "Exceção")
    tot_p += p; tot_a += a; tot_e += e
    for i, v in enumerate([f"{uc[0]} — {uc[1]}", p, a, e, p + a + e], 1):
        cell = ws4.cell(row=r, column=i, value=v)
        if i > 1:
            cell.alignment = center
    r += 1
aplica_bordas_zebra(ws4, hrow4 + 1, r - 1, len(cab4))
# linha total
for i, v in enumerate(["TOTAL", tot_p, tot_a, tot_e, tot_p + tot_a + tot_e], 1):
    cell = ws4.cell(row=r, column=i, value=v)
    cell.font = Font(bold=True)
    cell.fill = PatternFill("solid", fgColor=AZUL_CLARO)
    cell.border = border
    if i > 1:
        cell.alignment = center
for i, w in enumerate([40, 12, 14, 12, 10], 1):
    ws4.column_dimensions[get_column_letter(i)].width = w
ws4.freeze_panes = "A4"

out = "casos_de_uso.xlsx"
wb.save(out)
print(f"Planilha gerada: {out}")
print(f"Casos de uso: {len(CASOS)} | Cenários: {len(CENARIOS)}")
print(f"Por tipo -> Principal:{tot_p} Alternativo:{tot_a} Exceção:{tot_e}")
